"""
exports.py — Exportação de dados em Excel (.xlsx) e CSV
Usa openpyxl para Excel e csv nativo para CSV.
"""
import csv
import io
from datetime import datetime

from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    OPENPYXL = True
except ImportError:
    OPENPYXL = False


# ── Paleta da Igreja ───────────────────────────────────────────────────────────
GOLD   = 'C9A84C'
NAVY   = '1A2340'
WHITE  = 'FFFFFF'
LIGHT  = 'F5F3EE'
BORDER = 'DDD8CC'


def _xlsx_response(filename: str):
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _csv_response(filename: str):
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _style_header_row(ws, row: int, cols: int):
    """Estiliza a linha de cabeçalho com cor dourada."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = Border(
            bottom=Side(style='thin', color=GOLD),
            right=Side(style='thin', color=BORDER),
        )


def _style_data_row(ws, row: int, cols: int, even: bool):
    """Estiliza linha de dados com alternância."""
    bg = LIGHT if even else WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(vertical='center')
        cell.border    = Border(
            bottom=Side(style='thin', color=BORDER),
            right=Side(style='thin', color=BORDER),
        )


def _add_title_row(ws, title: str, cols: int):
    """Adiciona linha de título mesclada no topo."""
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(bold=True, color=WHITE, size=13)
    cell.fill      = PatternFill('solid', fgColor=GOLD)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28


def _auto_width(ws):
    """Ajusta largura das colunas automaticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)


def _now_str():
    return datetime.now().strftime('%Y%m%d_%H%M')


# ══════════════════════════════════════════════════════════
# ── 1. Membros ────────────────────────────────────────────
# ══════════════════════════════════════════════════════════

def export_membros_xlsx(queryset):
    from members.models import Presenca

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Membros'
    ws.freeze_panes = 'A3'

    headers = [
        'Nome completo', 'Username', 'E-mail', 'Telefone',
        'Grupo', 'Batizado', 'Membro desde', 'Nascimento',
        'Ministérios', 'Presenças', '% Frequência', 'Cadastro',
    ]

    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, profile in enumerate(queryset, start=1):
        pres_total   = Presenca.objects.filter(member=profile).count()
        pres_present = Presenca.objects.filter(member=profile, present=True).count()
        pct          = round(pres_present / pres_total * 100) if pres_total else 0
        ministerios  = ', '.join(
            profile.member_ministries.select_related('ministry')
                   .values_list('ministry__name', flat=True)
        )

        row = [
            profile.full_name,
            profile.user.username,
            profile.user.email,
            profile.phone or '—',
            profile.get_role_display(),
            'Sim' if profile.baptized else 'Não',
            profile.member_since.strftime('%d/%m/%Y') if profile.member_since else '—',
            profile.birth_date.strftime('%d/%m/%Y') if profile.birth_date else '—',
            ministerios or '—',
            pres_present,
            f'{pct}%',
            profile.created_at.strftime('%d/%m/%Y'),
        ]
        ws.append(row)
        _style_data_row(ws, i + 1, len(headers), i % 2 == 0)

    _add_title_row(ws, f'Lista de Membros — {datetime.now().strftime("%d/%m/%Y")}', len(headers))
    _auto_width(ws)
    ws.row_dimensions[2].height = 20

    resp = _xlsx_response(f'membros_{_now_str()}.xlsx')
    wb.save(resp)
    return resp


def export_membros_csv(queryset):
    resp = _csv_response(f'membros_{_now_str()}.csv')
    writer = csv.writer(resp)
    writer.writerow([
        'Nome', 'Username', 'E-mail', 'Telefone', 'Grupo',
        'Batizado', 'Membro desde', 'Nascimento', 'Ministérios', 'Cadastro',
    ])
    for profile in queryset:
        ministerios = ', '.join(
            profile.member_ministries.select_related('ministry')
                   .values_list('ministry__name', flat=True)
        )
        writer.writerow([
            profile.full_name,
            profile.user.username,
            profile.user.email,
            profile.phone or '',
            profile.get_role_display(),
            'Sim' if profile.baptized else 'Não',
            profile.member_since.strftime('%d/%m/%Y') if profile.member_since else '',
            profile.birth_date.strftime('%d/%m/%Y') if profile.birth_date else '',
            ministerios,
            profile.created_at.strftime('%d/%m/%Y'),
        ])
    return resp


# ══════════════════════════════════════════════════════════
# ── 2. Presenças ──────────────────────────────────────────
# ══════════════════════════════════════════════════════════

def export_presencas_xlsx(cultos_qs, members_qs):
    """
    Gera planilha cruzada: linhas = membros, colunas = cultos.
    Célula = ✓ (presente) ou · (ausente).
    """
    from members.models import Presenca

    cultos  = list(cultos_qs.order_by('date'))
    members = list(members_qs.order_by('user__first_name', 'user__last_name'))

    # Pré-carregar todas as presenças de uma vez
    pres_map = {}
    for p in Presenca.objects.filter(culto__in=cultos, member__in=members):
        pres_map[(p.culto_id, p.member_id)] = p.present

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Presenças'
    ws.freeze_panes = 'C3'

    # Cabeçalho: Membro | Total | % | culto1 | culto2 ...
    header = ['Membro', 'Presenças', '%'] + [
        f'{c.date.strftime("%d/%m")}\n{c.get_culto_type_display()[:12]}'
        for c in cultos
    ]
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    green_fill = PatternFill('solid', fgColor='D4EDDA')
    red_fill   = PatternFill('solid', fgColor='F8D7DA')

    for i, member in enumerate(members, start=1):
        pres_list = [pres_map.get((c.id, member.id), False) for c in cultos]
        total     = sum(pres_list)
        pct       = round(total / len(cultos) * 100) if cultos else 0

        row_data = [member.full_name, total, f'{pct}%'] + [
            '✓' if p else '·' for p in pres_list
        ]
        ws.append(row_data)
        row_num = i + 1
        _style_data_row(ws, row_num, len(header), i % 2 == 0)

        # Colorir células de presença
        for j, present in enumerate(pres_list, start=4):
            cell = ws.cell(row=row_num, column=j)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill      = green_fill if present else red_fill
            cell.font      = Font(
                color='155724' if present else '721C24',
                bold=present,
            )

    _add_title_row(
        ws,
        f'Relatório de Presenças — {datetime.now().strftime("%d/%m/%Y")}',
        len(header)
    )

    # Largura fixa para colunas de cultos
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    for col_idx in range(4, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    resp = _xlsx_response(f'presencas_{_now_str()}.xlsx')
    wb.save(resp)
    return resp


def export_presencas_csv(cultos_qs, members_qs):
    from members.models import Presenca

    cultos  = list(cultos_qs.order_by('date'))
    members = list(members_qs.order_by('user__first_name', 'user__last_name'))
    pres_map = {}
    for p in Presenca.objects.filter(culto__in=cultos, member__in=members):
        pres_map[(p.culto_id, p.member_id)] = p.present

    resp = _csv_response(f'presencas_{_now_str()}.csv')
    writer = csv.writer(resp)

    header = ['Membro', 'Presenças', '%'] + [
        f'{c.date.strftime("%d/%m/%Y")} - {c.get_culto_type_display()}'
        for c in cultos
    ]
    writer.writerow(header)

    for member in members:
        pres_list = [pres_map.get((c.id, member.id), False) for c in cultos]
        total     = sum(pres_list)
        pct       = round(total / len(cultos) * 100) if cultos else 0
        writer.writerow(
            [member.full_name, total, f'{pct}%'] +
            ['Sim' if p else 'Não' for p in pres_list]
        )
    return resp


# ══════════════════════════════════════════════════════════
# ── 3. Contribuições / Ofertas ────────────────────────────
# ══════════════════════════════════════════════════════════

def export_contribuicoes_xlsx(queryset):
    from decimal import Decimal

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contribuições'
    ws.freeze_panes = 'A3'

    headers = ['Membro', 'Tipo', 'Valor (R$)', 'Data', 'Observações', 'Registrado em']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    total = Decimal('0.00')
    for i, o in enumerate(queryset.order_by('-date'), start=1):
        total += o.amount
        row = [
            o.profile.full_name,
            o.get_type_display(),
            float(o.amount),
            o.date.strftime('%d/%m/%Y'),
            o.notes or '—',
            o.created_at.strftime('%d/%m/%Y'),
        ]
        ws.append(row)
        _style_data_row(ws, i + 1, len(headers), i % 2 == 0)
        # Formatar coluna de valor
        ws.cell(row=i + 1, column=3).number_format = 'R$ #,##0.00'

    # Linha de total
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=2, value='TOTAL')
    ws.cell(row=total_row, column=3, value=float(total))
    ws.cell(row=total_row, column=3).number_format = 'R$ #,##0.00'
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=GOLD)

    _add_title_row(ws, f'Contribuições — {datetime.now().strftime("%d/%m/%Y")}', len(headers))
    _auto_width(ws)

    resp = _xlsx_response(f'contribuicoes_{_now_str()}.xlsx')
    wb.save(resp)
    return resp


def export_contribuicoes_csv(queryset):
    resp = _csv_response(f'contribuicoes_{_now_str()}.csv')
    writer = csv.writer(resp)
    writer.writerow(['Membro', 'Tipo', 'Valor', 'Data', 'Observações'])
    for o in queryset.order_by('-date'):
        writer.writerow([
            o.profile.full_name,
            o.get_type_display(),
            str(o.amount),
            o.date.strftime('%d/%m/%Y'),
            o.notes or '',
        ])
    return resp


# ══════════════════════════════════════════════════════════
# ── 4. Inscrições em Eventos ──────────────────────────────
# ══════════════════════════════════════════════════════════

def export_inscricoes_xlsx(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inscrições'
    ws.freeze_panes = 'A3'

    headers = ['Evento', 'Data do Evento', 'Membro', 'E-mail', 'Status', 'Inscrito em']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, reg in enumerate(queryset.select_related('event', 'user').order_by('event__date'), 1):
        row = [
            reg.event.title,
            reg.event.date.strftime('%d/%m/%Y %H:%M') if reg.event.date else '—',
            reg.user.get_full_name() or reg.user.username,
            reg.user.email,
            reg.get_status_display() if hasattr(reg, 'get_status_display') else reg.status,
            reg.registered_at.strftime('%d/%m/%Y') if hasattr(reg, 'registered_at') else '—',
        ]
        ws.append(row)
        _style_data_row(ws, i + 1, len(headers), i % 2 == 0)

    _add_title_row(ws, f'Inscrições em Eventos — {datetime.now().strftime("%d/%m/%Y")}', len(headers))
    _auto_width(ws)

    resp = _xlsx_response(f'inscricoes_{_now_str()}.xlsx')
    wb.save(resp)
    return resp


# ══════════════════════════════════════════════════════════
# ── 5. Visitantes ─────────────────────────────────────────
# ══════════════════════════════════════════════════════════

def export_visitantes_xlsx(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Visitantes'
    ws.freeze_panes = 'A3'

    headers = ['Nome', 'E-mail', 'Telefone', 'Como conheceu', 'Contactado', 'Data']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, v in enumerate(queryset.order_by('-created_at'), start=1):
        row = [
            v.name,
            v.email or '—',
            v.phone or '—',
            v.how_did_you_know or '—',
            'Sim' if v.contacted else 'Não',
            v.created_at.strftime('%d/%m/%Y'),
        ]
        ws.append(row)
        _style_data_row(ws, i + 1, len(headers), i % 2 == 0)

    _add_title_row(ws, f'Visitantes — {datetime.now().strftime("%d/%m/%Y")}', len(headers))
    _auto_width(ws)

    resp = _xlsx_response(f'visitantes_{_now_str()}.xlsx')
    wb.save(resp)
    return resp


def export_visitantes_csv(queryset):
    resp = _csv_response(f'visitantes_{_now_str()}.csv')
    writer = csv.writer(resp)
    writer.writerow(['Nome', 'E-mail', 'Telefone', 'Como conheceu', 'Contactado', 'Data'])
    for v in queryset.order_by('-created_at'):
        writer.writerow([
            v.name,
            v.email or '',
            v.phone or '',
            v.how_did_you_know or '',
            'Sim' if v.contacted else 'Não',
            v.created_at.strftime('%d/%m/%Y'),
        ])
    return resp


# ══════════════════════════════════════════════════════════
# ── 6. Aniversariantes ────────────────────────────────────
# ══════════════════════════════════════════════════════════

def export_aniversariantes_xlsx(queryset):
    from datetime import date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Aniversariantes'
    ws.freeze_panes = 'A3'

    headers = ['Nome', 'E-mail', 'Telefone', 'Data de Nascimento', 'Mês', 'Idade']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    hoje = date.today()
    for i, profile in enumerate(queryset, start=1):
        bd = profile.birth_date
        idade = hoje.year - bd.year - ((hoje.month, hoje.day) < (bd.month, bd.day)) if bd else '—'
        row = [
            profile.full_name,
            profile.user.email,
            profile.phone or '—',
            bd.strftime('%d/%m/%Y') if bd else '—',
            bd.strftime('%B') if bd else '—',
            idade,
        ]
        ws.append(row)
        _style_data_row(ws, i + 1, len(headers), i % 2 == 0)

    _add_title_row(ws, f'Aniversariantes — {datetime.now().strftime("%d/%m/%Y")}', len(headers))
    _auto_width(ws)

    resp = _xlsx_response(f'aniversariantes_{_now_str()}.xlsx')
    wb.save(resp)
    return resp
